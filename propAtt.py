import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
import numpy as np
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

def save_bar_chart_with_labels(group_data, title, sheet_name, workbook):
    # Drop any columns or rows with all NaN values
    group_data = group_data.dropna(axis=1, how='all')
    group_data = group_data.dropna(axis=0, how='all')

    # Check if the DataFrame is empty after dropping NaN values
    if group_data.empty:
        print(f"No data to plot for {sheet_name}. Skipping chart generation.")
        return
    
    # Convert decimals to percentages for plotting
    group_data = group_data * 100

    # Adjust positions array to match the number of valid data points
    positions = np.arange(len(group_data.index))  # Only for valid rows
    

    fig, ax = plt.subplots(figsize=(16, 8))
    bar_width = 0.8 / len(group_data.columns)  # Calculate width of each bar with space between them
    
    # Define colors for the bars
    colors = plt.colormaps['tab10']
    
    for i, col in enumerate(group_data.columns):  # Include all remaining columns
        grade_label = col  # Directly use the grade as an integer
        
        bars = ax.bar(positions + i * bar_width, group_data[col].values, width=bar_width, label=str(grade_label), color=colors(i))
        for j, bar in enumerate(bars):
            height = bar.get_height()
            if not np.isnan(height):
                ax.text(bar.get_x() + bar.get_width() / 2, height - 1, f'{grade_label}: {height:.2f}%', ha='center', va='top', fontsize=10, rotation=90, color='black')
    
    ax.set_xticks(positions + bar_width * (len(group_data.columns) - 1) / 2)
    ax.set_xticklabels(group_data.index, rotation=45, ha='right')
    ax.set_title(title)
    ax.set_ylabel('Attendance Rate (%)')
  
    ax.set_ylim(0, 100)
    y_ticks = np.arange(0, 101, 10)  # Dynamic ticks based on the minimum value
    
    ax.set_yticks(y_ticks)


    ax.legend(title="Grade Levels")
    plt.tight_layout()
    
    # Save the figure temporarily
    chart_path = f"{sheet_name}.png"
    plt.savefig(chart_path)
    plt.close(fig)
    
    # Add the chart to the Excel sheet
    img = Image(chart_path)
    ws = workbook[sheet_name]
    ws.add_image(img, 'A1')  # Add the image to the worksheet


def prepare_data(file_path, school_dict, school_type_dict, exclude_schools=None):
    # Read the CSV file
    df = pd.read_csv(file_path)
    
    # Map school codes to names and types
    df['School_Type'] = df['Reporting School'].map(school_type_dict)
    df['School Name'] = df['Reporting School'].map(school_dict)
    
    # Filter out unnecessary rows
    if exclude_schools:
        df = df[~df['Reporting School'].isin(exclude_schools)]
   
    return df

def pull_alt_specprog_data(df, altspecprog_mapping):
    # Filter out the Webster Learning Center and specprog data
    alt_specprog_data = df[df['Attending School'].isin(altspecprog_mapping)]
    
    # Map the school codes to names
    alt_specprog_data['School Name'] = alt_specprog_data['Attending School'].map(altspecprog_mapping)
    
    return alt_specprog_data

def create_alt_specprog_pivot_table(df, values, index, columns=None, aggfunc='mean', margins=False, margins_name='Total'):
    # Ensure data types are appropriate
    if 'Grade' in df.columns:
        df['Grade'] = df['Grade'].astype(int)  # Ensure 'Grade' is an integer before pivoting

    # Create the pivot table
    pivot_table = pd.pivot_table(df, 
                                 values=values, 
                                 index=index, 
                                 columns=columns, 
                                 aggfunc=aggfunc,
                                 margins=margins, 
                                 margins_name=margins_name)
    
    # Make sure the Grade index is kept as integer
    if 'Grade' in pivot_table.index.names:
       pivot_table.index = pivot_table.index.astype(int)

    return pivot_table

def calculate_attendance_rates(df):
    # District-wide calculations
    df['Sum_Adj_Prop_Wt_Grade'] = df.groupby(['Grade'])['Adj Prop Wt'].transform('sum')
    df['Sum_Segment_Grade'] = df.groupby(['Grade'])['Segment'].transform('sum')

    df['Attendance Rate DistrictWide'] = df['Sum_Adj_Prop_Wt_Grade'] / df['Sum_Segment_Grade'].replace(0, np.nan)

    # Building-level calculations
    df['Sum_Adj_Prop_Wt_School'] = df.groupby(['Reporting School', 'Grade'])['Adj Prop Wt'].transform('sum')
    df['Sum_Segment_School'] = df.groupby(['Reporting School', 'Grade'])['Segment'].transform('sum')

    df['Attendance Rate School'] = df['Sum_Adj_Prop_Wt_School'] / df['Sum_Segment_School'].replace(0, np.nan)
    
    return df

def create_pivot_table(df, values, index, columns=None, aggfunc='mean', margins=False, margins_name='Total'):
    if 'Grade' in df.columns:
        df['Grade'] = df['Grade'].astype(int)  # Ensure 'Grade' is integer before pivoting

    pivot_table = pd.pivot_table(df, 
                                 values=values, 
                                 index=index, 
                                 columns=columns, 
                                 aggfunc=aggfunc,
                                 margins=margins, 
                                 margins_name=margins_name)
    if 'Grade' in pivot_table.index.names:
       pivot_table.index = pivot_table.index.astype(int)  # Ensure 'Grade' stays integer in the pivot table

    return pivot_table


# Specific for District-wide by Grade to include weighted points and total possible points
def create_district_wide_pivot(df):
    pivot_table = pd.pivot_table(df, 
                                 values=['Attendance Rate DistrictWide', 'Sum_Adj_Prop_Wt_Grade', 'Sum_Segment_Grade'], 
                                 index=['Grade'],
                                 aggfunc='mean')
    return pivot_table

def write_to_excel(excel_path, pivot_tables, df):
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        for sheet_name, pivot_table in pivot_tables.items():
            pivot_table.to_excel(writer, sheet_name=sheet_name)
        df.to_excel(writer, sheet_name='Raw Data')

def apply_formatting(wb, sheet_name, percent_columns, autosize=True):
    ws = wb[sheet_name]
    
    # Apply the percentage number format directly
    percent_format = '0.00%'

    # Apply the percentage format directly to the relevant columns
    for col in percent_columns:
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{get_column_letter(col)}{row}"]
            cell.number_format = percent_format

    # Auto-size columns if needed
    if autosize:
        autosize_columns(ws)


def write_combined_pivot_tables(wb, sheet_name, pivot_tables):
    ws = wb.create_sheet(sheet_name)
    
    # Insert the title row with the current date and time
    title = f"Weighted Attendance as of {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.cell(row=1, column=1, value=title)

    start_row = 3  # Start at the first row
    table_ranges = {}

    for table_name, pivot_table in pivot_tables.items():
        # Convert column headers to strings
        pivot_table.columns = pivot_table.columns.astype(str)
        
        # Write the header row
        for col_num, header in enumerate(pivot_table.columns.insert(0, pivot_table.index.name), 1):
            ws.cell(row=start_row, column=col_num, value=header)

        # Write the pivot table data
        for r_idx, row in enumerate(pivot_table.itertuples(index=True, name=None), start=start_row + 1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Record the start and end rows for this table
        end_row = start_row + len(pivot_table)
        table_ranges[table_name] = (start_row + 1, end_row)  # +1 to skip the header

        # Convert the written range into a table
        start_col_letter = get_column_letter(1)
        end_col_letter = get_column_letter(len(pivot_table.columns) + 1)
        table_range = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
        
        table_name_str = f"Table_{table_name.replace(' ', '_')}"
        table = Table(displayName=table_name_str, ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style

        # Check if the table name already exists in the worksheet
        existing_table_names = [t for t in ws._tables]
        if table_name_str not in existing_table_names:
            ws.add_table(table)

        # Update the starting row for the next pivot table
        start_row = end_row + 2  # +3 to add one empty row between tables and account for the header

    return table_ranges

# Auto-size columns for worksheet
def autosize_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = [cell.value for cell in col]
        for cell in column:
            try:
                if len(str(cell)) > max_length:
                    max_length = len(str(cell))
            except:
                pass

        # Set custom width for columns containing percentages
        if ws.title == 'Building Level by Grade' and col[0].column >= 2 and col[0].column <= 14:
            max_length = 5  # Set custom width to accommodate percentage
        
        # New block to auto-size columns B to N (2 to 14) for 'alt and specprog' sheet
        if ws.title == 'alt and specprog' and col[0].column >= 2 and col[0].column <= 14:
            # Adjust max_length if needed for specific columns
            max_length = 5  # Ensure at least a reasonable width

        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

def write_district_wide_table(wb, pivot_table, sheet_name):
    # Add or access the sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Define the start position
    start_row = 1
    start_col = 1

    # Write the headers
    for col_num, header in enumerate(pivot_table.columns.insert(0, pivot_table.index.name), start_col):
        ws.cell(row=start_row, column=col_num, value=header)

    # Write the data
    for r_idx, row in enumerate(pivot_table.itertuples(index=True, name=None), start=start_row + 1):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Define the table range
    end_row = start_row + len(pivot_table)
    end_col = start_col + len(pivot_table.columns)
    table_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

    # Generate a unique table name based on a timestamp or counter
    import time
    unique_suffix = str(int(time.time()))  # Use the current time as a unique suffix
    table_name = f"Table_DistrictWide_{unique_suffix}"

    # Create and style the table
    table = Table(displayName=table_name, ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

def create_district_wide_graph(wb, pivot_table, sheet_name):
    # Create a new sheet for the graph
    wb.create_sheet(sheet_name)

    # Prepare data for the graph
    fig, ax = plt.subplots(figsize=(16, 8))
    bar_width = 0.8  # Width of each bar
    positions = np.arange(len(pivot_table))  # Positions for the bars
    
    # Define a colormap for the bars
    colors = plt.colormaps['tab10'](np.linspace(0, 1, len(pivot_table)))

    # Create the bars
    bars = ax.bar(positions, pivot_table['Attendance Rate DistrictWide'] * 100, width=bar_width, color=colors)

    # Add labels inside the bars
    for i, bar in enumerate(bars):
        height = bar.get_height()  # Use the actual height of the bar (which is now correct)
        grade = int(pivot_table.index[i])  # Grade value as an integer
        ax.text(
            bar.get_x() + bar.get_width() / 2, height - 5,  # Place the label slightly below the top of the bar
            f'{grade}\n{height:.2f}%', 
            ha='center', va='top', fontsize=10, rotation=90, color='white'  # Use white color for visibility inside the bar
        )
    
    ax.set_xticks(positions)
    ax.set_xticklabels(pivot_table.index.astype(int), rotation=45, ha='right')
    ax.set_title('District-wide Attendance Rate by Grade')
    ax.set_ylabel('Attendance Rate (%)')
    ax.set_ylim(0, 100)  # Ensure the y-axis is set to 100% to match the expected percentage scale
    plt.tight_layout()

    # Save the figure temporarily
    chart_path = f"{sheet_name}.png"
    plt.savefig(chart_path)
    plt.close(fig)

    # Add the chart to the Excel sheet
    img = Image(chart_path)
    ws = wb[sheet_name]
    ws.add_image(img, 'A1')  # Add the image to the worksheet

def generate_graph_sheets(wb, school_pivot_tables):
    for sheet_name, pivot_table in school_pivot_tables.items():
        if sheet_name == "Elementary Schools":
            # Split elementary schools into two halves
            midpoint = len(pivot_table) // 2
            first_half = pivot_table.iloc[:midpoint]
            second_half = pivot_table.iloc[midpoint:]
            
            # Create first graph for the first half
            first_half_sheet_name = f"{sheet_name} Graph Part 1"
            wb.create_sheet(first_half_sheet_name)
            save_bar_chart_with_labels(first_half, f"{sheet_name} Attendance Rate by Grade - Part 1", first_half_sheet_name, wb)
            
            # Create second graph for the second half
            second_half_sheet_name = f"{sheet_name} Graph Part 2"
            wb.create_sheet(second_half_sheet_name)
            save_bar_chart_with_labels(second_half, f"{sheet_name} Attendance Rate by Grade - Part 2", second_half_sheet_name, wb)
        else:
            graph_sheet_name = f"{sheet_name} Graph"
            wb.create_sheet(graph_sheet_name)
            save_bar_chart_with_labels(pivot_table, f"{sheet_name} Attendance Rate by Grade", graph_sheet_name, wb)


def apply_formatting_to_range(ws, start_row, end_row, percent_columns):
    percent_format = '0.00%'
    for col in percent_columns:
        for row in range(start_row, end_row + 1):
            cell = ws[f"{get_column_letter(col)}{row}"]
            cell.number_format = percent_format

def remove_existing_table(ws, table_name):
    # Remove the table with the given name, if it exists
    ws._tables = {table for table in ws._tables if table.name != table_name}

def write_alt_specprog_table(wb, pivot_table, sheet_name):
    # Add or access the sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Define the start position
    start_row = 1
    start_col = 1

    # Write the headers
    for col_num, header in enumerate(pivot_table.columns.insert(0, pivot_table.index.name), start_col):
        ws.cell(row=start_row, column=col_num, value=header)

    # Write the data
    for r_idx, row in enumerate(pivot_table.itertuples(index=True, name=None), start=start_row + 1):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Define the table range
    end_row = start_row + len(pivot_table)
    end_col = start_col + len(pivot_table.columns)
    table_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

    # Create and style the table
    table = Table(displayName="Table_alt_specprog", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)



def main():
    {
        #high schools
        1: 'HS 1',
        2: 'HS 2',
        3: 'HS 3',
        #middle schools
        4: 'MS 1',
        5: 'MS 2',
        6: 'MS 3',
        7: 'MS 4',
        #elementary schools
        8: 'ES 1',
        9: 'ES 2',
        10: 'ES 3',
        11: 'ES 4',
        12: 'ES 5',
        13: 'ES 6',
        14: 'ES 7',
        15: 'ES 8',
        16: 'ES 9',
        17: 'ES 10',
        18: 'ES 11',
        19: 'ES 12',
        20: 'ES 13',
        #early learning centers
        21: 'ELC 1',
        22: 'ELC 2',
    }

    
    school_type_dict = {
        21: 'ELC',
        22: 'ELC',
        1: 'High School',
        2: 'High School',
        3: 'High School',
        4: 'Middle School',
        5: 'Middle School',
        6: 'Middle School',
        7: 'Middle School',
        8: 'Elementary School',
        9: 'Elementary School',
        10: 'Elementary School',
        11: 'Elementary School',
        12: 'Elementary School',
        13: 'Elementary School',
        14: 'Elementary School',
        15: 'Elementary School',
        16: 'Elementary School',
        17: 'Elementary School',
        18: 'Elementary School',
        19: 'Elementary School',
        20: 'Elementary School'
    }

    altspecprog_mapping = {
        6000: 'alt Elem',
        98: 'alt MS',
        99: 'alt HS',
        6001: 'specprog Elem',
        6002: 'specprog MS',
        6003: 'specprog HS'
    }
    
    exclude_schools = [23, 24]

    # File paths and dictionary data
    file_path = 'YTD Student Attendance Extract.csv'
    # Generate the Excel file name with the current date
    current_date = datetime.now().strftime('%Y-%m-%d_%H-%M')
    excel_path = f'Attendance_Report_{current_date}.xlsx'

    # Delete the existing file if it exists
    if os.path.exists(excel_path):
        os.remove(excel_path)

    # Now, create a new workbook
    wb = Workbook()
    
    wb.save(excel_path)

    # Load the workbook
    wb = load_workbook(excel_path)

    # Step 1: Prepare the data
    df = prepare_data(file_path, school_dict, school_type_dict, exclude_schools)

    # Step 2: Calculate attendance rates
    df = calculate_attendance_rates(df)
    
    # Step 3: Create pivot tables
    district_wide_pivot = create_district_wide_pivot(df)

    write_district_wide_table(wb, district_wide_pivot, 'District-wide by Grade')

    # Remove the default blank sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    create_district_wide_graph(wb, district_wide_pivot, 'District-wide Graph')

    school_pivot_tables = {
        'High Schools': create_pivot_table(df[df['School_Type'] == 'High School'], values='Attendance Rate School', index=['School Name'], columns=['Grade'], margins=True),
        'Middle Schools': create_pivot_table(df[df['School_Type'] == 'Middle School'], values='Attendance Rate School', index=['School Name'], columns=['Grade'], margins=True),
        #'Elementary Schools': create_pivot_table(df[df['School_Type'] == 'Elementary School'], values='Attendance Rate School', index=['School Name'], columns=['Grade'], margins=True),
        'Elementary Schools': create_pivot_table(df[(df['School_Type'] == 'Elementary School') & (df['Grade'] != -1)],  # Exclude grade -1 for Elementary Schools
                                             values='Attendance Rate School', index=['School Name'], columns=['Grade'], margins=True),
        'Early Learning Centers': create_pivot_table(df[df['School_Type'] == 'ELC'], values='Attendance Rate School', index=['School Name'], columns=['Grade'], margins=True)
    }

    # Step 5: Write combined pivot tables (excluding District-wide by Grade) to the same sheet
    table_ranges = write_combined_pivot_tables(wb, 'Combined Schools', school_pivot_tables)
    
    # Call the function to generate graph sheets   
    generate_graph_sheets(wb, school_pivot_tables)
        
    # Step 6: Apply formatting
    apply_formatting(wb, 'District-wide by Grade', percent_columns=[2])
       
    # Apply formatting for Combined Schools based on tracked ranges
    ws_combined = wb['Combined Schools']
    for table_name, (start_row, end_row) in table_ranges.items():
        if table_name == 'High Schools':
            apply_formatting_to_range(ws_combined, start_row, end_row, percent_columns=range(2, 7))
        elif table_name == 'Middle Schools':
            apply_formatting_to_range(ws_combined, start_row, end_row, percent_columns=range(2, 6))
        elif table_name == 'Elementary Schools':
            apply_formatting_to_range(ws_combined, start_row, end_row, percent_columns=range(2, 10))
        elif table_name == 'Early Learning Centers':
            apply_formatting_to_range(ws_combined, start_row, end_row, percent_columns=range(2, 6))
    # Add the raw data to the workbook directly
    if 'Raw Data' not in wb.sheetnames:
        ws_raw_data = wb.create_sheet('Raw Data')
    else:
        ws_raw_data = wb['Raw Data']

    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws_raw_data.cell(row=r_idx, column=c_idx, value=value)

    # Optionally, you can also add headers
    for col_num, header in enumerate(df.columns, start=1):
        ws_raw_data.cell(row=1, column=col_num, value=header)
    
    #*******************************new alt specprog Portion********************************
    # New Steps for alt and specprog Data:
    
    # Step 7: Filter alt and specprog data
    alt_specprog_data = pull_alt_specprog_data(df, altspecprog_mapping)

    # Step 8: Create pivot table for alt and specprog
    alt_specprog_pivot = create_alt_specprog_pivot_table(alt_specprog_data, 
                                             values='Attendance Rate School', 
                                             index=['School Name'], 
                                             columns=['Grade'], 
                                             margins=True)

    # Step 9: Write alt and specprog pivot table to a new sheet
    write_alt_specprog_table(wb, alt_specprog_pivot, 'alt and specprog')

    # Step 10: Create a graph for alt and specprog data
    #create_district_wide_graph(wb, alt_specprog_pivot, 'alt and specprog Graph')

    # Step 11: Apply formatting to alt and specprog sheet
    apply_formatting(wb, 'alt and specprog', percent_columns=range(2, 15))
    #*******************************new alt specprog Portion********************************

    # Save the workbook
    wb.save(excel_path)

# Entry point
if __name__ == "__main__":
    main()